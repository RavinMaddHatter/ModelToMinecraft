import trimesh
import numpy
import amulet
from amulet.api.block import Block
import os
import shutil
from openpyxl import Workbook
from zipfile import ZipFile 
from tkinter import DoubleVar,StringVar,IntVar, Label, Entry, Tk, Button
from tkinter import filedialog
import threading
import LevelDatNbt as worldNBT
from trimesh.transformations import euler_matrix
import time
import ntpath

YOffset=4
desiredSize=50
buildCenterX=0
buildCenterY=0
root = Tk()
root.title("Madhatter's Minecraft Model Maker")
FileGUI=StringVar()
FileGUI.set("creeper.stl")
BaseGUI=StringVar()
BaseGUI.set("Base World.mcworld")
sizeGui=IntVar()
sizeGui.set(desiredSize)
startHGui=IntVar()
startHGui.set(YOffset)
buildCenterXGui=IntVar()
buildCenterXGui.set(0)
buildCenterYGui=IntVar()
buildCenterYGui.set(0)

rotXGui=DoubleVar()
rotYGui=DoubleVar()
rotZGui=DoubleVar()
rotXGui.set(0)
rotYGui.set(0)
rotZGui.set(0)

outputGui=StringVar()
outputGui.set("")

def browseObj():
    FileGUI.set(filedialog.askopenfilename(filetypes = (("3D Model", "*.stl *.STL"), )))
def browseBase():
    BaseGUI.set(filedialog.askopenfilename(filetypes = (("Template files", "*.mcworld"), )))
def run():
    outputGui.set("starting")
    fileName=FileGUI.get()
    YOffset=startHGui.get()
    desiredSize=sizeGui.get()
    buildCenterX=buildCenterXGui.get()
    print(buildCenterX)
    buildCenterY=buildCenterYGui.get()
    print(buildCenterY)
    templateWorld=BaseGUI.get()
    ## setup parameters
    outputFileName=fileName.replace(".STL","").replace(".stl","").replace(".obj","").replace(".OBJ","")
    path_to_save="temp"
    blocks=0
    solidBlocks=0
    wb = Workbook()
    xrot=rotXGui.get()
    yrot=rotYGui.get()
    zrot=rotZGui.get()

    ##Make import mesh and turn it into a big
    m = trimesh.load(fileName) #import a mesh
    HTM=a=euler_matrix(xrot*numpy.pi/180,yrot*numpy.pi/180,zrot*numpy.pi/180)
    m.apply_transform(HTM)
    biggest=max(m.bounding_box.primitive.extents)
    outputGui.set("sizing...")
    v=m.voxelized(pitch=biggest/desiredSize)#scale to match biggest dimension.
    dims=v.matrix.shape
    outV=v.matrix.copy()
    outputGui.set("voxel magic done")

    ##make temp folder
    if not os.path.isdir(path_to_save):#make a temp folder to work out of
        os.mkdir(path_to_save)
    else:
        for filename in os.listdir(path_to_save):
            file_path = os.path.join(path_to_save, filename)
            try:
                if os.path.isfile(file_path) or os.path.islink(file_path):
                    os.unlink(file_path)
                elif os.path.isdir(file_path):
                    shutil.rmtree(file_path)
            except Exception as e:
                print('Failed to delete %s. Reason: %s' % (file_path, e))
    with ZipFile(templateWorld, 'r') as zipObj:
        zipObj.extractall(path_to_save)
    
    outputGui.set("hollowing the model")
    ##Processing the world to make it hollow and put the excel sheet together
    world = amulet.load_level(path_to_save)
    selectedBlock = Block("minecraft", "smooth_stone")
    game_version = ("bedrock", (1, 16, 210))  # the version that we want the block data in.
    
    for z in range(dims[2]):
        ws1 = wb.create_sheet("Y layer " +str(z+YOffset))
        for y in range(dims[1]):
            for x in range(dims[0]):
                if x>0 and y>0 and z>0:
                    if x+1<dims[0] and y+1<dims[1] and z+1<dims[2]:
                        if v.matrix[x+1][y][z] and v.matrix[x-1][y][z]:
                            if v.matrix[x][y+1][z] and v.matrix[x][y-1][z]:
                                if v.matrix[x][y][z+1] and v.matrix[x][y][z-1]:
                                    outV[x][y][z]=False
                if outV[x][y][z]:
                    blocks+=1
                    world.set_version_block(
                        x-round(dims[0]/2),  # x location
                        z+YOffset,  # y location
                        y-round(dims[2]/2),  # z location
                        "minecraft:overworld",  # dimension
                        game_version,
                        selectedBlock,
                        )
                    ws1.cell(row=y+1,column=x+1).value=str(-x+round(dims[0]/2)+buildCenterX)+"/"+str(-y+round(dims[2]/2)+buildCenterY)
                if v.matrix[x][y][z]:
                    solidBlocks+=1
    world.save()
    world.close()
        
    with worldNBT.BedrockLevelFile.load(os.path.join(path_to_save,"level.dat")) as lvlNBT:
        lvlNBT["LastPlayed"]=time.time()
        lvlNBT["LevelName"]=ntpath.basename(outputFileName)
    wb.save(filename = outputFileName+'.xlsx')
    wb.close()
    outputGui.set("saving")

    #Clean up temp files
    startDir=os.getcwd()
    os.chdir(os.path.join(startDir,path_to_save))
    with ZipFile(outputFileName+".mcworld", 'w') as zipF:
        zipF.write("world_icon.jpeg")
        zipF.write("levelname.txt")
        zipF.write("level.dat_old")
        zipF.write("level.dat")
        for root, dirs, files in os.walk("db"):
            for file in files:
                zipF.write(os.path.join(root, file))
    ##pack up world 
    shutil.move(os.path.join(os.getcwd(),outputFileName+".mcworld"),os.path.join(startDir,outputFileName+".mcworld"))
    #print blocks required.
    outputGui.set("numblocks:%d num stacks:%f num shulkers:%f"%(blocks,blocks/64,blocks/64/27))


def runThread():
    x = threading.Thread(target=run)
    x.start()

runbutton=Button(root, text="Make My world",command=runThread)
getFileButton=Button(root, text="Browse model",command=browseObj)
getBaseButton=Button(root, text="Browse Base",command=browseBase)

fileLB=Label(root, text="3D file")
BaseLB=Label(root, text="Minecraft World")
fileEntry = Entry(root,textvariable=FileGUI)
BaseEntry = Entry(root,textvariable=BaseGUI)

maxSizeLB=Label(root, text="Max Size in Blocks")
maxDimEntry = Entry(root,textvariable=sizeGui)

BuildXLB=Label(root, text="Build Center X (for excel only)")
BuildXEntry = Entry(root,textvariable=buildCenterXGui)
BuildYLB=Label(root, text="Build Center Z (for excel only)")
BuildYEntry = Entry(root,textvariable=buildCenterYGui)
OutputEntry = Entry(root,textvariable=outputGui,width=44)

rotXLabel=Label(root, text="Rotation About the X Axis")
rotYLabel=Label(root, text="Rotation About the Y Axis")
rotZLabel=Label(root, text="Rotation About the Z Axis")
rotXEntry = Entry(root,textvariable=rotXGui)
rotYEntry = Entry(root,textvariable=rotYGui)
rotZEntry = Entry(root,textvariable=rotZGui)

HeightOffGround= Entry(root,textvariable=startHGui)
HeightOffGroundLB=Label(root, text="Starting Y Level")
r=0
fileLB.grid(row=r,column=0)
fileEntry.grid(row=r,column=1)
getFileButton.grid(row=r,column=2)
r+=1
BaseEntry.grid(row=r,column=1)
BaseLB.grid(row=r,column=0)

getBaseButton.grid(row=r,column=2)
r+=1
maxSizeLB.grid(row=r,column=0)
maxDimEntry.grid(row=r,column=1)
r+=1
HeightOffGround.grid(row=r,column=1)
HeightOffGroundLB.grid(row=r,column=0)
r+=1
rotXLabel.grid(row=r,column=0)
rotXEntry.grid(row=r,column=1)
r+=1
rotYLabel.grid(row=r,column=0)
rotYEntry.grid(row=r,column=1)
r+=1
rotZLabel.grid(row=r,column=0)
rotZEntry.grid(row=r,column=1)
r+=1
BuildXEntry.grid(row=r,column=1)
BuildXLB.grid(row=r,column=0)
r+=1
BuildYEntry.grid(row=r,column=1)
BuildYLB.grid(row=r,column=0)
r+=1


OutputEntry.grid(row=r,column=0,columnspan = 2)
r+=1
runbutton.grid(row=r,column=2)
root.mainloop() 


